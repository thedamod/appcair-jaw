"""Extraction pipeline: 687 documents -> a serializable knowledge graph (KG).

Stage 1 of the system. Parses 678 PDFs (certificates, reference letters,
personnel records, CVs) with config-driven label/value scanning plus prose
fallback regexes, merges the two certificate sides of each work by normalized
work name, indexes clients and engineers, and reduces the 9 Excel workbooks to
per-client financial facts. Everything is cached behind a corpus fingerprint
so the KG is rebuilt only when the corpus actually changes.
"""

import hashlib
import json
import os
import re
import glob

import fitz

ROOT = os.path.dirname(os.path.abspath(__file__))
DOCS = os.path.join(ROOT, "BITS-Hackathon-Dataset", "documents")
CACHE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cache")
CONFIG = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config")
os.makedirs(CACHE, exist_ok=True)


def _load_config():
    with open(os.path.join(CONFIG, "fields.json")) as fh:
        return json.load(fh)


CFG = _load_config()
_MONTHS = {k.lower(): int(v) for k, v in CFG["months"].items()}
_UNITS = [(u, int(f)) for u, f in CFG["money_units"]]
_CLIENT_SUFFIX = re.compile(CFG["client_suffix_regex"])

_LABEL_RE = re.compile(r"^[A-Z][A-Za-z /&'()-]{2,62}?:?$")
_SECTION_RE = re.compile(r"^\d{1,2}\.\s")
_PAGE_RE = re.compile(r"(?i)^page\s+\d+\s+of")
_MAX_LABEL_LEN = 64
_MAX_VALUE_LEN = 200

_FLAG_MAP = {"i": re.I, "m": re.M, "s": re.S, "im": re.I | re.M, "is": re.I | re.S}
_DOC_TYPES = {}
for _dt, _d in CFG["doc_types"].items():
    _label_to_field = {}
    _parse_of_field = {}
    for _f in _d["fields"]:
        _parse_of_field[_f["field"]] = _f["parse"]
        for _lab in _f["labels"]:
            _label_to_field[_lab.strip().lower().rstrip(":")] = _f["field"]
    _prose = []
    for _p in _d.get("prose", []):
        _prose.append({
            "field": _p["field"],
            "post": _p.get("post"),
            "re": re.compile(_p["regex"], _FLAG_MAP.get(_p.get("flags", ""), 0)),
            "group": _p.get("group", 1),
        })
    _DOC_TYPES[_dt] = {
        "label_to_field": _label_to_field,
        "parse_of_field": _parse_of_field,
        "client_header_lines": int(_d.get("client_header_lines", 0)),
        "prose": _prose,
    }

_GRADE_PATTERNS = []
for _p in CFG["grades"]["patterns"]:
    _GRADE_PATTERNS.append({
        "re": re.compile(_p["regex"], _FLAG_MAP.get(_p.get("flags", ""), 0)),
        "group": _p.get("group"),
        "value": _p.get("value"),
    })

_PCERT = CFG["personnel"]
_CV = CFG["cv"]
_FIN = CFG["financial"]


def parse_money(s):
    """Parse any Indian rendering of rupees ('INR 33.38 Cr', '3,338.00 Lakh',
    '33,38,00,000', 'Rs. 5,00,000/-') into a lossless integer of rupees."""
    if not s:
        return None
    t = (s.replace("INR", " ").replace("Rs.", " ").replace("Rs ", " ")
         .replace("₹", " ").replace("/-", "").strip())
    low = t.lower()
    mult = None
    for unit, factor in _UNITS:
        if re.search(r"(?i)\b" + unit + r"\b", low):
            mult = factor
            t = re.sub(r"(?i)\b" + unit + r"\b", "", t)
            break
    t = t.replace(",", "").strip()
    if not re.search(r"\d", t):
        return None
    try:
        return int(round(float(t) * mult)) if mult else int(round(float(t)))
    except ValueError:
        return None


def _dt(y, mo, d):
    try:
        return (int(y), int(mo), int(d))
    except (TypeError, ValueError):
        return None


def parse_date(text):
    """Parse a date in any of the corpus's dialects (ISO, day-first, "d Mon yyyy",
    "Mon d, yyyy") into a JSON-safe (year, month, day) tuple."""
    if not text:
        return None
    m = re.search(r"\b(20\d{2})[-/.](\d{1,2})[-/.](\d{1,2})\b", text)
    if m:
        return _dt(m.group(1), m.group(2), m.group(3))
    m = re.search(r"\b(\d{1,2})[-/.](\d{1,2})[-/.](\d{4})\b", text)
    if m and int(m.group(3)) > 1000:
        return _dt(m.group(3), m.group(2), m.group(1))
    m = re.search(r"\b(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]{3,9})[,]?\s+(20\d{2})\b", text)
    if m:
        mo = _MONTHS.get(m.group(2).lower()[:3])
        return _dt(m.group(3), mo, m.group(1))
    m = re.search(r"\b([A-Za-z]{3,9})\s+(\d{1,2})(?:st|nd|rd|th)?[,]?\s+(20\d{2})\b", text)
    if m:
        mo = _MONTHS.get(m.group(1).lower()[:3])
        return _dt(m.group(3), mo, m.group(2))
    return None


def days_between(a, b):
    try:
        import datetime
        return (datetime.date(*b) - datetime.date(*a)).days
    except Exception:
        return None


def client_canonical(raw):
    """Normalize a client name: strip the trailing '(...)' suffix and collapse
    whitespace so the same client is one key everywhere in the KG."""
    if not raw:
        return None
    n = _CLIENT_SUFFIX.sub("", raw.strip())
    n = re.sub(r"\s+", " ", n)
    return n.strip().rstrip(".").strip()


def _clean(s):
    return re.sub(r"\s+", " ", s or "").strip()


def pkg_of(work):
    m = re.search(r"(?i)Pkg[- ]?(\d+)", work or "")
    return m.group(1) if m else None


def state_of(work):
    m = re.search(r"(?i)—\s*([A-Za-z ]+?)\s+Pkg", work or "")
    return m.group(1).strip() if m else None


def extract_table_pairs(text):
    """Scan document text for (label, value) pairs in either corpus layout:
    a label line followed by its value line, or an inline 'Label: value' line.
    Position-independent on purpose - 62 issuers format certificates differently."""
    lines = text.split("\n")
    n = len(lines)
    pairs = []

    i = 0
    while i < n:
        line = lines[i].strip()
        if 3 < len(line) <= _MAX_LABEL_LEN and _LABEL_RE.match(line):
            j = i + 1
            while j < n and not lines[j].strip():
                j += 1
            if j < n:
                value = lines[j].strip()
                if (value and 0 < len(value) <= _MAX_VALUE_LEN
                        and not _SECTION_RE.match(value)
                        and not _PAGE_RE.match(value)):
                    pairs.append((line.rstrip(":").strip(), value))
                    i = j + 1
                    continue
        m = re.match(r"^([A-Z][A-Za-z /&'()-]{2,62}?)\s*[:：]\s*(.{2,200})$", line)
        if m:
            pairs.append((m.group(1).strip(), m.group(2).strip()))
        i += 1

    return pairs


def resolve_fields(pairs, doc_cfg):
    out = {}
    seen = set()
    l2f = doc_cfg["label_to_field"]
    for label, value in pairs:
        key = label.lower()
        field = l2f.get(key)
        if field and field not in out and key not in seen:
            out[field] = value
            seen.add(key)
    return out


_WORK_QUOTED = re.compile(r'["“”]\s*([^"“”]{3,}?(?:—|–|-)\s*[A-Za-z]+(?:\s+[A-Za-z]+)*\s*Pkg-?\d+[^"“”]*?)\s*["“”]')


def extract_work_name(text):
    m = _WORK_QUOTED.search(text)
    if m:
        return m.group(1).strip()
    return None


def parse_grade(text):
    for p in _GRADE_PATTERNS:
        m = p["re"].search(text)
        if not m:
            continue
        if p.get("value") is not None:
            return p["value"]
        return m.group(p["group"]).lower()
    return None


_PARSERS = {
    "text": lambda v: v.strip(),
    "money": parse_money,
    "date": parse_date,
    "role": lambda v: v.strip().lower(),
}


def parse_doc(doc_type, text):
    """Parse one document of a given type into a record: table pairs first,
    prose fallback regexes second, client letterhead and grade last."""
    doc_cfg = _DOC_TYPES[doc_type]
    lines = [l for l in text.split("\n") if l.strip()]
    rec = {"work": extract_work_name(text)}

    fields = resolve_fields(extract_table_pairs(text), doc_cfg)
    for field, raw in fields.items():
        parse = doc_cfg["parse_of_field"].get(field)
        if parse == "work_name":
            if re.search(r"(?i)Pkg", raw):
                rec["work"] = raw.strip()
        elif parse in _PARSERS:
            rec[field] = _PARSERS[parse](raw)

    for p in doc_cfg["prose"]:
        if p["field"] in rec:
            continue
        m = p["re"].search(text)
        if not m:
            continue
        raw = m.group(p["group"])
        rec[p["field"]] = _PARSERS[p["post"]](raw) if p["post"] in _PARSERS else raw.strip()

    if doc_cfg["client_header_lines"]:
        head = lines[:doc_cfg["client_header_lines"]]
        rec["client_raw"] = _clean(head[0]) if head else None

    if doc_type in ("completion_certificate", "company_completion_certificate"):
        rec["grade"] = parse_grade(text)

    return rec


def parse_pcert(text):
    rec = {}
    m = re.search(_PCERT["certify_regex"], text)
    if m:
        rec["name"] = m.group(1).strip()
    else:
        m = re.search(_PCERT["confer_regex"], text)
        rec["name"] = m.group(1).strip() if m else None
    rec["cred_type"] = None
    for lab in _PCERT["cred_type_labels"]:
        fm = re.search(r"(?i)\b" + re.escape(lab) + r"\s*[:：]?\s*\n?\s*([^\n]+)", text)
        if fm:
            rec["cred_type"] = fm.group(1).strip()
            break
    cid = re.search(_PCERT["cred_id_regex"], text)
    rec["cred_id"] = cid.group(1) if cid else None
    rec["issue_date"] = parse_date(text)
    return rec


def parse_cv(text):
    m = re.search(_CV["name_regex"], text)
    return {"name": m.group(1).strip() if m else None}


def doc_text(path):
    d = fitz.open(path)
    try:
        return "\n".join(p.get_text() for p in d)
    finally:
        d.close()


_ALL_DOC_DIRS = ["completion_certificate", "company_completion_certificate",
                 "reference_letter", "personnel_certificate", "cv"]


def _wkey(n):
    return re.sub(r"\s+", " ", (n or "").strip()).lower()


def build(verbose=True):
    """Assemble the knowledge graph from the whole corpus.

    Merge strategy: company completion certificates (155, carry the client)
    seed each work keyed by normalized quoted work name; client completion
    certificates (155) fill value/date/pm/grade onto the same keys; reference
    letters (132) mark works as referenced and add the contractor role.
    """
    works = {}
    cc_files = sorted(glob.glob(os.path.join(DOCS, "completion_certificate", "*.pdf")))
    ccc_files = sorted(glob.glob(os.path.join(DOCS, "company_completion_certificate", "*.pdf")))
    ref_files = sorted(glob.glob(os.path.join(DOCS, "reference_letter", "*.pdf")))

    for f in ccc_files:
        rec = parse_doc("company_completion_certificate", doc_text(f))
        k = _wkey(rec["work"])
        if not k:
            continue
        w = works.setdefault(k, {"work": rec["work"]})
        w["client"] = client_canonical(rec.get("client") or rec.get("client_raw"))
        if rec.get("category"):
            w["category"] = rec["category"]
        if rec.get("value") is not None:
            w["value"] = rec["value"]
        if rec.get("completion_date"):
            w["completion_date"] = rec["completion_date"]
        if rec.get("pm"):
            w["pm"] = rec["pm"]
        if rec.get("grade") and not w.get("grade"):
            w["grade"] = rec["grade"]

    for f in cc_files:
        rec = parse_doc("completion_certificate", doc_text(f))
        k = _wkey(rec["work"])
        if not k:
            continue
        w = works.setdefault(k, {"work": rec["work"]})
        if rec.get("category"):
            w["category"] = rec["category"]
        if rec.get("value") is not None:
            w["value"] = rec["value"]
        if rec.get("completion_date"):
            w["completion_date"] = rec["completion_date"]
        if rec.get("pm"):
            w["pm"] = rec["pm"]
        if rec.get("grade"):
            w["grade"] = rec["grade"]

    known = list(works.keys())
    for f in ref_files:
        txt = doc_text(f)
        rec = parse_doc("reference_letter", txt)
        hit = None
        if rec.get("work") and _wkey(rec["work"]) in works:
            hit = _wkey(rec["work"])
        else:
            # PDF extraction may wrap a state name across lines; normalize
            # whitespace before matching the canonical work key.
            low = re.sub(r"\s+", " ", txt).lower()
            for k in known:
                if k in low:
                    hit = k
                    break
        if hit:
            works[hit]["referenced"] = True
            if rec.get("role"):
                works[hit]["role"] = rec["role"]
            if rec.get("category") and not works[hit].get("category"):
                works[hit]["category"] = rec["category"]

    clients, engineers = {}, {}
    for k, w in works.items():
        c = w.get("client")
        if c:
            clients.setdefault(client_canonical(c), set()).add(k)
        pm = w.get("pm")
        if pm:
            engineers.setdefault(pm.strip(), set()).add(k)
        w["pkg"] = pkg_of(w["work"])
        w["state"] = state_of(w["work"])

    people = set()
    for f in sorted(glob.glob(os.path.join(DOCS, "personnel_certificate", "*.pdf")) +
                    glob.glob(os.path.join(DOCS, "cv", "*.pdf"))):
        txt = doc_text(f)
        rec = parse_pcert(txt)
        if rec.get("name"):
            people.add(rec["name"].strip())
        rec = parse_cv(txt)
        if rec.get("name"):
            people.add(rec["name"].strip())
    people |= set(engineers.keys())

    if verbose:
        print("works:", len(works), "| clients:", len(clients), "| engineers:", len(engineers),
              "| people:", len(people),
              "| referenced:", sum(1 for w in works.values() if w.get("referenced")))
        print("works without value:", sum(1 for w in works.values() if w.get("value") is None))
        print("works without client:", sum(1 for w in works.values() if not w.get("client")))
        print("works without date:", sum(1 for w in works.values() if not w.get("completion_date")))
        print("works without grade:", sum(1 for w in works.values() if not w.get("grade")))
        print("works without pm:", sum(1 for w in works.values() if not w.get("pm")))
    return {"works": works, "clients": clients, "engineers": engineers, "people": people}


def serialize(kg):
    return {"works": {k: v for k, v in kg["works"].items()},
            "clients": {k: sorted(v) for k, v in kg["clients"].items()},
            "engineers": {k: sorted(v) for k, v in kg["engineers"].items()},
            "people": sorted(kg["people"])}


def corpus_fingerprint():
    h = hashlib.sha256()
    for f in sorted(glob.glob(os.path.join(DOCS, "**", "*.*"), recursive=True)):
        st = os.stat(f)
        h.update(f"{f}:{st.st_mtime_ns}:{st.st_size}".encode())
    return h.hexdigest()[:16]


def ensure_kg(verbose=False):
    """Return the KG, rebuilding from the corpus only if its fingerprint changed."""
    kg_path = os.path.join(CACHE, "kg.json")
    fp = corpus_fingerprint()
    if os.path.exists(kg_path):
        try:
            meta = json.load(open(kg_path))
            if meta.get("fingerprint") == fp:
                return meta["kg"]
        except (json.JSONDecodeError, KeyError):
            pass
    kg = build(verbose=verbose)
    with open(kg_path, "w") as fh:
        json.dump({"fingerprint": fp, "kg": serialize(kg)}, fh, indent=1)
    return kg


def _col_index(header, name):
    for i, h in enumerate(header):
        if name in h or h in name:
            return i
    return None


def build_financial(verbose=True):
    """Reduce the 9 workbooks to per-sheet totals and per-client column sums."""
    import openpyxl
    facts = {}
    for f in sorted(glob.glob(os.path.join(DOCS, "workbooks", "*.xlsx"))):
        name = os.path.basename(f).replace(".xlsx", "")
        rec = {"sheets": {}}
        try:
            wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
        except Exception as e:
            rec["error"] = str(e)
            facts[name] = rec
            continue
        for s in wb.sheetnames:
            ws = wb[s]
            rows = [r for r in ws.iter_rows(values_only=True)
                    if any(c is not None for c in r)]
            if not rows:
                continue
            header = [str(c).lower() if c is not None else "" for c in rows[0]]
            data = rows[1:]
            totals = {}
            for col in _FIN["column_vocab"]:
                idx = _col_index(header, col)
                if idx is None:
                    continue
                vals = []
                for r in data:
                    try:
                        v = r[idx]
                        if isinstance(v, (int, float)):
                            vals.append(float(v))
                    except Exception:
                        pass
                if vals:
                    totals[col] = round(sum(vals), 2)
            by_client = {}
            cidx = _col_index(header, "client")
            if cidx is not None:
                for r in data:
                    try:
                        cl = client_canonical(str(r[cidx])) if r[cidx] else None
                    except Exception:
                        cl = None
                    if not cl:
                        continue
                    g = by_client.setdefault(cl, {"count": 0})
                    g["count"] += 1
                    for col in _FIN["client_columns"]:
                        idx = _col_index(header, col)
                        if idx is None:
                            continue
                        try:
                            if isinstance(r[idx], (int, float)):
                                g[col] = g.get(col, 0) + float(r[idx])
                        except Exception:
                            pass
                for cl in by_client:
                    by_client[cl] = {k: round(v, 2) if isinstance(v, float) else v
                                     for k, v in by_client[cl].items()}
            rec["sheets"][s] = {"rows": len(data), "totals": totals,
                                "by_client": by_client}
        wb.close()
        facts[name] = rec
    if verbose:
        for k, v in facts.items():
            print(" workbook:", k, {s: v["sheets"][s]["totals"] for s in v.get("sheets", {})})
    return facts


if __name__ == "__main__":
    kg = build()
    with open(os.path.join(CACHE, "kg.json"), "w") as fh:
        json.dump({"fingerprint": corpus_fingerprint(), "kg": serialize(kg)}, fh, indent=1)
